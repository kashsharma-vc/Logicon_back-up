import os
from datetime import datetime
from io import BytesIO
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

def generate_attendance_excel(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"

    # Define styles
    header_fill = PatternFill(start_color="0A1128", end_color="0A1128", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=20, bold=True, color="000000")
    subtitle_font = Font(size=10, italic=True, color="666666")
    column_header_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Row 1-2: Title and Subtitle
    ws.merge_cells('B1:J1')
    ws['B1'] = "LOGICON FIELD OPERATIONS"
    ws['B1'].font = title_font
    ws['B1'].alignment = left_align

    ws.merge_cells('B2:J2')
    generated_date = timezone.localdate().strftime("%B %d, %Y")
    ws['B2'] = f"User Activity & Attendance Report • Generated {generated_date}"
    ws['B2'].font = subtitle_font
    ws['B2'].alignment = left_align

    # Add Logo
    logo_path = r'C:\field-senses-app-main\Main Logicon\FE-Logicon-Connect-ATS-main\public\logo_temp.png'
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 100
        img.height = 30
        ws.add_image(img, 'A1')

    # Row 4: DAILY SESSION AUDIT LOGS
    ws.merge_cells('A4:J4')
    ws['A4'] = "DAILY SESSION AUDIT LOGS"
    ws['A4'].fill = header_fill
    ws['A4'].font = header_font
    ws['A4'].alignment = left_align

    # Row 6: SESSION METRICS SUMMARY
    ws.merge_cells('A6:D6')
    ws['A6'] = "SESSION METRICS SUMMARY"
    ws['A6'].font = bold_font

    active_sessions = sum(1 for session in qs if session.status == 'active')
    total_logs = len(qs)

    ws['A7'] = "Active Sessions"
    ws['A7'].font = bold_font
    ws['B7'] = active_sessions
    ws['C7'] = "Total Logs"
    ws['C7'].font = bold_font
    ws['D7'] = total_logs

    # Row 9: Table Headers
    headers = [
        "Employee Name", "Employee Code", "Role", "Department",
        "Login Time", "Logout Time", "Duration", "Session Status",
        "Attendance Status", "IP Address"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.fill = column_header_fill
        cell.alignment = center_align

    # Data Rows
    row_num = 10
    for session in qs:
        emp = session.employee
        emp_name = emp.get_full_name() or emp.email
        emp_code = emp.employee_code if hasattr(emp, 'employee_code') and emp.employee_code else 'N/A'
        role = emp.user_type.capitalize() if emp.user_type else '-'
        dept = emp.department.name if hasattr(emp, 'department') and emp.department else 'N/A'
        
        login_time = session.check_in_at.strftime('%Y-%m-%d %H:%M:%S') if session.check_in_at else '-'
        logout_time = session.check_out_at.strftime('%Y-%m-%d %H:%M:%S') if session.check_out_at else 'Active Now'
        
        hours = int(session.total_hours or 0)
        minutes = int(((session.total_hours or 0) - hours) * 60)
        duration = f"{hours}h {minutes}m" if (session.total_hours or 0) > 0 else "0h 0m"
        
        session_status = 'ACTIVE' if session.status == 'active' else 'COMPLETED'
        
        # Attendance Status Logic
        attendance_status = "PRESENT"
        if session.check_in_at and session.check_in_at.strftime('%H:%M') > '09:30':
            attendance_status = "LATE"
        elif session.status != 'active' and (session.total_hours or 0) < 8:
            attendance_status = "UNDER_HOURS"
            
        ip_address = session.check_in_ip if hasattr(session, 'check_in_ip') and session.check_in_ip else '127.0.0.1'

        row_data = [
            emp_name, emp_code, role, dept,
            login_time, logout_time, duration,
            session_status, attendance_status, ip_address
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num > 4:
                cell.alignment = center_align
                
        row_num += 1

    # Adjust Column Widths
    column_widths = {
        'A': 25, 'B': 15, 'C': 15, 'D': 20,
        'E': 20, 'F': 20, 'G': 12, 'H': 15,
        'I': 18, 'J': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()
