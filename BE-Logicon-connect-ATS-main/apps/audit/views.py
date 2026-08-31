from rest_framework import viewsets, mixins, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import AuditLog, UserActivityLog, EmailReportSettings
from .serializers import AuditLogSerializer, UserActivityLogSerializer, EmailReportSettingsSerializer

from apps.access.viewsets import ScopedReadOnlyModelViewSet
from apps.access.querysets import filter_user_activity_logs_for_user
from apps.access.permissions import HasCapability


class AuditLogViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """Read-only audit log — no create/update/delete via API."""
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['actor', 'org', 'object_type', 'action']
    search_fields = ['action', 'object_type', 'object_id']


def generate_activity_logs_workbook(queryset, user):
    """Generates a beautifully styled openpyxl Workbook matching Logicon template rules."""
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone

    # Create a workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activity & Attendance Logs"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Define styles
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_header = Font(name="Calibri", size=10, bold=True, color="1E293B")
    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Try to insert Logo
    logo_inserted = False
    try:
        from openpyxl.drawing.image import Image as OpenPyxlImage
        logo_path = r"c:\field-senses-app-main\Main Logicon\FE-Logicon-Connect-ATS-main\public\LOGO-2-1.webp"
        if os.path.exists(logo_path):
            from PIL import Image as PILImage
            img = PILImage.open(logo_path)
            temp_png_path = os.path.join(os.path.dirname(logo_path), "logo_temp.png")
            img.save(temp_png_path, "PNG")

            ox_img = OpenPyxlImage(temp_png_path)
            ox_img.width = 45
            ox_img.height = 45
            ws.add_image(ox_img, "A1")
            logo_inserted = True
    except Exception as e:
        print("Failed to add image:", e)

    # Title Block
    if logo_inserted:
        ws.merge_cells("B1:J2")
        cell = ws["B1"]
        cell.value = "LOGICON FIELD OPERATIONS"
        cell.font = Font(name="Calibri", size=18, bold=True, color="0F172A")
        cell.alignment = Alignment(vertical="center")

        ws.merge_cells("B3:J3")
        ws["B3"] = f"User Activity & Attendance Report • Generated {timezone.localtime(timezone.now()).strftime('%B %d, %Y')}"
        ws["B3"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
    else:
        ws.merge_cells("A1:J2")
        cell = ws["A1"]
        cell.value = "LOGICON FIELD OPERATIONS"
        cell.font = Font(name="Calibri", size=18, bold=True, color="0F172A")
        cell.alignment = Alignment(vertical="center", horizontal="center")

        ws.merge_cells("A3:J3")
        ws["A3"] = f"User Activity & Attendance Report • Generated {timezone.localtime(timezone.now()).strftime('%B %d, %Y')}"
        ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
        ws["A3"].alignment = Alignment(horizontal="center")

    # Spacing row
    ws.row_dimensions[4].height = 15

    # Dark Navy Strip Header
    ws.merge_cells("A5:J5")
    strip_cell = ws["A5"]
    strip_cell.value = "  DAILY SESSION AUDIT LOGS"
    strip_cell.font = font_title
    strip_cell.fill = navy_fill
    strip_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[5].height = 25

    # spacing
    ws.row_dimensions[6].height = 10

    # Section Summary / KPIs Block
    ws.merge_cells("A7:B7")
    ws["A7"] = "SESSION METRICS SUMMARY"
    ws["A7"].font = font_section

    # Counters row
    ws["A8"] = "Active Sessions"
    ws["A8"].font = font_bold
    ws["A8"].fill = light_blue_fill
    ws["A8"].border = thin_border

    active_count = queryset.filter(session_status='active').count()
    ws["B8"] = active_count
    ws["B8"].font = font_data
    ws["B8"].border = thin_border

    ws["C8"] = "Total Logs"
    ws["C8"].font = font_bold
    ws["C8"].fill = light_blue_fill
    ws["C8"].border = thin_border

    ws["D8"] = queryset.count()
    ws["D8"].font = font_data
    ws["D8"].border = thin_border

    ws.row_dimensions[8].height = 20

    # spacing
    ws.row_dimensions[9].height = 15

    # Headers for Main Data
    headers = [
        "Employee Name", "Employee Code", "Role", "Department",
        "Login Time", "Logout Time", "Duration", "Session Status",
        "Attendance Status", "IP Address"
    ]

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = text
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[11].height = 25

    # Populate Data
    current_row = 12
    for log in queryset:
        name = "N/A"
        code = "N/A"
        role = ""
        dept = "N/A"
        
        if log.user:
            name = f"{log.user.first_name} {log.user.last_name}".strip() or log.user.username
            code = log.user.employee_code or "N/A"
            if log.user.department:
                dept = log.user.department.name or "N/A"
            try:
                assignment = log.user.role_assignments.first()
                if assignment and assignment.role:
                    role = assignment.role.name or ""
            except Exception:
                pass

        login = "N/A"
        if log.login_time:
            try:
                login = timezone.localtime(log.login_time).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                login = str(log.login_time)

        logout = "Active Now"
        if log.logout_time:
            try:
                logout = timezone.localtime(log.logout_time).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                logout = str(log.logout_time)

        # Duration
        duration = "Active Now"
        if log.logout_time and log.login_time:
            diff = log.logout_time - log.login_time
            hours = int(diff.total_seconds() // 3600)
            minutes = int((diff.total_seconds() % 3600) // 60)
            duration = f"{hours}h {minutes}m"

        row_data = [
            name, code, role, dept,
            login, logout, duration,
            (log.session_status or 'ACTIVE').upper(),
            (log.attendance_status or 'PRESENT').upper(),
            log.ip_address or "N/A"
        ]

        fill = zebra_fill if current_row % 2 == 0 else white_fill
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = font_data
            cell.fill = fill
            cell.border = thin_border
            if col_idx in [1, 2, 3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        first_cell = col[0]
        col_letter = get_column_letter(first_cell.column) if isinstance(first_cell.column, int) else str(first_cell.column)
        for cell in col:
            if cell.row < 11:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return wb


def send_daily_log_report_email(user, subject, body):
    """Generates the excel workbook for user scope daily logs and emails it to them."""
    import io
    from django.core.mail import EmailMessage
    from django.utils import timezone
    from apps.access.querysets import filter_user_activity_logs_for_user

    # Fetch logs scoped for the user
    queryset = UserActivityLog.objects.select_related('user', 'user__department').order_by('-login_time')
    queryset = filter_user_activity_logs_for_user(queryset, user)

    # Filter to today's logs only
    today_start = timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    queryset = queryset.filter(login_time__gte=today_start)

    # Generate workbook bytes
    wb = generate_activity_logs_workbook(queryset, user)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Send mail
    email = EmailMessage(
        subject=subject or "Daily Attendance & Session Logs Report",
        body=body or "Please find attached the daily log report.",
        from_email=None,
        to=[user.email],
    )

    filename = f"Logicon_Activity_Report_{timezone.localtime(timezone.now()).strftime('%Y%m%d')}.xlsx"
    email.attach(filename, buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    email.send()


class UserActivityLogViewSet(ScopedReadOnlyModelViewSet):
    queryset = UserActivityLog.objects.select_related('user', 'user__department').order_by('-login_time')
    serializer_class = UserActivityLogSerializer
    permission_classes = [IsAuthenticated, HasCapability]
    scope_filter = filter_user_activity_logs_for_user
    filterset_fields = ['session_status', 'attendance_status']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'user__employee_code']

    def get_required_capability(self):
        return 'user.read'

    def get_queryset(self):
        qs = super().get_queryset()
        date_str = self.request.query_params.get('date')
        if date_str:
            try:
                from django.utils.dateparse import parse_date
                parsed = parse_date(date_str)
                if parsed:
                    qs = qs.filter(login_time__date=parsed)
            except Exception:
                pass
        start_date = self.request.query_params.get('start_date')
        if start_date:
            try:
                from django.utils.dateparse import parse_date
                parsed = parse_date(start_date)
                if parsed:
                    qs = qs.filter(login_time__date__gte=parsed)
            except Exception:
                pass
        end_date = self.request.query_params.get('end_date')
        if end_date:
            try:
                from django.utils.dateparse import parse_date
                parsed = parse_date(end_date)
                if parsed:
                    qs = qs.filter(login_time__date__lte=parsed)
            except Exception:
                pass
        return qs

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        total_sessions = queryset.count()
        active_sessions = queryset.filter(session_status='active').count()

        has_date_filter = any(k in request.query_params for k in ['date', 'start_date', 'end_date'])
        if has_date_filter:
            target_logs = queryset
        else:
            from django.utils import timezone
            today_start = timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
            target_logs = queryset.filter(login_time__gte=today_start)

        present_today = target_logs.filter(attendance_status__in=['present', 'late', 'under_hours']).values('user').distinct().count()
        late_today = target_logs.filter(attendance_status='late').values('user').distinct().count()

        dept_breakdown = {}
        for log in target_logs:
            dept = log.user.department.name if (log.user and log.user.department) else 'Unassigned'
            dept_breakdown[dept] = dept_breakdown.get(dept, 0) + 1

        from apps.access.models import UserRoleAssignment
        role_breakdown = {}
        user_ids = target_logs.values_list('user_id', flat=True).distinct()
        assignments = UserRoleAssignment.objects.filter(user_id__in=user_ids).select_related('role')
        user_to_role = {a.user_id: a.role.name for a in assignments if a.role}

        for log in target_logs:
            r = user_to_role.get(log.user_id, 'No Role')
            role_breakdown[r] = role_breakdown.get(r, 0) + 1

        device_breakdown = {'Desktop': 0, 'Mobile': 0, 'Tablet': 0}
        for log in queryset[:500]:
            ua = (log.user_agent or '').lower()
            if 'ipad' in ua or 'tablet' in ua:
                device_breakdown['Tablet'] += 1
            elif 'mobile' in ua or 'android' in ua or 'iphone' in ua:
                device_breakdown['Mobile'] += 1
            else:
                device_breakdown['Desktop'] += 1

        return Response({
            'kpis': {
                'total_sessions': total_sessions,
                'active_now': active_sessions,
                'present_today': present_today,
                'late_today': late_today,
            },
            'department_breakdown': dept_breakdown,
            'role_breakdown': role_breakdown,
            'device_breakdown': device_breakdown
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        from django.http import HttpResponse
        from django.utils import timezone
        try:
            queryset = self.filter_queryset(self.get_queryset())
            wb = generate_activity_logs_workbook(queryset, request.user)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Logicon_Activity_Logs_{timezone.localtime(timezone.now()).strftime("%Y%m%d")}.xlsx'
            wb.save(response)
            return response
        except Exception as view_error:
            return HttpResponse(f"Error generating Excel report: {str(view_error)}", status=500, content_type="text/plain")


class EmailReportSettingsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        settings, created = EmailReportSettings.objects.get_or_create(user=request.user)
        serializer = EmailReportSettingsSerializer(settings)
        return Response(serializer.data)

    def create(self, request):
        settings, created = EmailReportSettings.objects.get_or_create(user=request.user)
        serializer = EmailReportSettingsSerializer(settings, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def test_email(self, request):
        settings, created = EmailReportSettings.objects.get_or_create(user=request.user)
        if not request.user.email:
            return Response({'detail': 'Your account does not have a configured email address.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            send_daily_log_report_email(request.user, settings.subject, settings.body)
            return Response({'detail': f'Test Excel report email sent to {request.user.email}.'})
        except Exception as email_err:
            return Response({'detail': f'Failed to send email: {str(email_err)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
