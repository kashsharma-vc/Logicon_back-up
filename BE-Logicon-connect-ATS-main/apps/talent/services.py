"""
apps/talent/services.py

Talent processing service helpers (Phase Talent-Hiring-A/B).
These are synchronous stubs — async/Celery/OCR/LLM dispatch is wired in a
later phase.  The status transitions are the authoritative state machine.
"""

import csv
import hashlib
import re
import uuid
from io import BytesIO, StringIO
from decimal import Decimal, InvalidOperation

from django.db import transaction
from rest_framework import serializers as drf_serializers
from rest_framework.exceptions import ValidationError

from .models import Resume


def normalize_phone(value: str) -> str:
    """
    Normalize Indian mobile number to a 10-digit string.
    Strips leading +91 / 91 country code if present.
    Raises DRF ValidationError if the result is not a valid 10-digit mobile.
    """
    value = value.strip().replace(' ', '').replace('-', '')
    if value.startswith('+91'):
        value = value[3:]
    elif value.startswith('91') and len(value) == 12:
        value = value[2:]
    if not re.match(r'^[6-9]\d{9}$', value):
        raise drf_serializers.ValidationError(
            "Enter a valid 10-digit Indian mobile number."
        )
    return value


def normalize_skill_name(skill: str) -> str:
    return skill.strip().lower()


def compute_file_hash(f) -> str:
    """Compute SHA-256 hex digest of an uploaded file. Resets file pointer after reading."""
    h = hashlib.sha256()
    if hasattr(f, 'seek'):
        f.seek(0)
    content = f.read() if hasattr(f, 'read') else b''
    h.update(content)
    if hasattr(f, 'seek'):
        f.seek(0)
    return h.hexdigest()


def determine_document_type(filename: str = '', content_type: str = '') -> str:
    """Return the normalized document type used by resume/import filters."""
    filename = (filename or '').lower().strip()
    content_type = (content_type or '').lower().strip()
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''

    if ext in {'pdf', 'docx', 'doc', 'txt', 'xlsx', 'csv'}:
        return ext
    if content_type == 'application/pdf':
        return 'pdf'
    if content_type == 'text/plain':
        return 'txt'
    if content_type in {
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    }:
        return 'docx'
    if content_type == 'application/msword':
        return 'doc'
    if content_type in {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
    }:
        return 'xlsx'
    if content_type in {'text/csv', 'application/csv'}:
        return 'csv'
    return 'unknown'


def queue_resume_processing(resume: Resume) -> None:
    """
    Check for duplicate file, then schedule background parsing via Celery.
    Sets status='duplicate_file' and returns early if an indexed resume with
    the same file_hash already exists.
    """
    if resume.file_hash:
        duplicate = (
            Resume.objects
            .filter(file_hash=resume.file_hash, status='indexed')
            .exclude(pk=resume.pk)
            .first()
        )
        if duplicate:
            Resume.objects.filter(pk=resume.pk).update(status='duplicate_file')
            resume.status = 'duplicate_file'
            return

    Resume.objects.filter(pk=resume.pk).update(status='extracting')
    resume.status = 'extracting'

    from apps.talent.tasks import process_resume_task
    process_resume_task.delay(resume.pk)


def mark_resume_manual_review(resume: Resume, reason: str) -> None:
    """Flag a resume for manual review with a reason string."""
    Resume.objects.filter(pk=resume.pk).update(
        status='manual_review',
        manual_review_reason=reason,
    )
    resume.status = 'manual_review'
    resume.manual_review_reason = reason


def mark_resume_failed(resume: Resume, error: str) -> None:
    """Mark a resume as failed and record the error message."""
    Resume.objects.filter(pk=resume.pk).update(
        status='failed',
        error_message=error,
    )
    resume.status = 'failed'
    resume.error_message = error


def build_candidate_profile_text(candidate) -> str:
    """
    Return a plain-text summary of a candidate's profile for search indexing
    or LLM context.  Pulls from latest indexed resume + skills + experience.
    """
    parts = [candidate.full_name]
    if candidate.current_role:
        parts.append(candidate.current_role)
    if candidate.current_company:
        parts.append(f"at {candidate.current_company}")
    if candidate.total_experience_years is not None:
        parts.append(f"{candidate.total_experience_years}y exp")

    skills = list(
        candidate.skills.values_list('skill_name', flat=True).order_by('skill_name')
    )
    if skills:
        parts.append("Skills: " + ", ".join(skills))

    for exp in candidate.experiences.order_by('-start_date')[:5]:
        line = f"{exp.job_title} @ {exp.company_name}"
        if exp.start_date:
            line += f" ({exp.start_date.year}"
            if exp.end_date:
                line += f"–{exp.end_date.year}"
            elif exp.is_current:
                line += "–present"
            line += ")"
        parts.append(line)

    return " | ".join(p for p in parts if p)


# Profile fields that can be updated on get_or_create (existing candidate).
# Only non-empty incoming values are applied.
_CANDIDATE_UPDATABLE_FIELDS = [
    'first_name', 'last_name', 'middle_name', 'email',
    'current_role', 'current_location', 'total_experience_years',
    'preferred_location', 'notice_period_days', 'current_company',
    'expected_ctc', 'current_ctc', 'target_job_role',
    'collar_type', 'source',
]


def manual_resume_intake(user, validated_data: dict) -> dict:
    """
    Create/update candidate, upload resume, tag skills, optionally create
    hiring application — all inside one atomic transaction.

    Returns dict: {candidate, resume, skills, hiring_application}
    """
    from rest_framework.exceptions import ValidationError

    from .models import Candidate, CandidateSkill

    with transaction.atomic():
        phone = validated_data['phone']
        phone_normalized = normalize_phone(phone)
        org = user.org

        # ── 1. Candidate get_or_create ─────────────────────────────────────
        defaults = {
            'phone': phone,
            'first_name': validated_data['first_name'],
            'last_name': validated_data['last_name'],
            'middle_name': validated_data.get('middle_name') or '',
            'email': validated_data.get('email') or '',
            'current_role': validated_data.get('current_role') or '',
            'current_location': validated_data.get('current_location') or '',
            'total_experience_years': validated_data.get('total_experience_years'),
            'preferred_location': validated_data.get('preferred_location') or '',
            'notice_period_days': validated_data.get('notice_period_days'),
            'current_company': validated_data.get('current_company') or '',
            'expected_ctc': validated_data.get('expected_ctc'),
            'current_ctc': validated_data.get('current_ctc'),
            'collar_type': validated_data.get('collar_type') or '',
            'source': validated_data.get('source') or 'manual',
            'billing_type': validated_data.get('billing_type'),
        }
        candidate, created = Candidate.objects.get_or_create(
            org=org,
            phone_normalized=phone_normalized,
            defaults=defaults,
        )

        # ── 2. Update existing candidate — never overwrite with blank ──────
        if not created:
            update_fields = []
            for field in _CANDIDATE_UPDATABLE_FIELDS:
                incoming = validated_data.get(field)
                if incoming is None or incoming == '':
                    continue
                if getattr(candidate, field) != incoming:
                    setattr(candidate, field, incoming)
                    update_fields.append(field)
            if validated_data.get('billing_type') and candidate.billing_type != validated_data['billing_type']:
                candidate.billing_type = validated_data['billing_type']
                update_fields.append('billing_type')
            if update_fields:
                candidate.save(update_fields=update_fields)

        # ── 3. Resume ──────────────────────────────────────────────────────
        f = validated_data['resume_file']
        resume = Resume.objects.create(
            candidate=candidate,
            file=f,
            original_filename=getattr(f, 'name', ''),
            content_type=getattr(f, 'content_type', ''),
            size_bytes=getattr(f, 'size', 0),
            document_type=determine_document_type(
                getattr(f, 'name', ''),
                getattr(f, 'content_type', ''),
            ),
            source_type='recruiter_upload',
            status='uploaded',
            uploaded_by=user,
            file_hash=compute_file_hash(f),
            view_only_note=validated_data.get('view_only_note') or '',
        )

        # ── 4. Skills — idempotent by normalized name ──────────────────────
        skills_out = []
        for skill_name in validated_data.get('skills') or []:
            normalized = normalize_skill_name(skill_name)
            skill, _ = CandidateSkill.objects.get_or_create(
                candidate=candidate,
                normalized_skill_name=normalized,
                defaults={
                    'skill_name': skill_name,
                    'source': 'manual',
                    'source_resume': resume,
                },
            )
            skills_out.append(skill)

        # ── 5. Optional hiring application ─────────────────────────────────
        hiring_app = None
        mrf_li = validated_data.get('mrf_line_item')
        mrf = validated_data.get('mrf')

        if mrf_li:
            from apps.hiring.models import (
                HiringApplication, ApplicationStageHistory, PipelineStage,
            )

            if candidate.org_id != mrf.org_id:
                raise ValidationError(
                    {'candidate': 'Candidate org does not match MRF org.'}
                )

            if HiringApplication.objects.filter(
                candidate=candidate, mrf_line_item=mrf_li,
            ).exists():
                raise ValidationError(
                    {'mrf_line_item': 'Candidate already has an application for this line item.'}
                )

            current_stage = validated_data.get('current_stage')
            if current_stage is None:
                current_stage = (
                    PipelineStage.objects
                    .filter(org=mrf.org, is_active=True)
                    .order_by('order')
                    .first()
                )

            hiring_app = HiringApplication.objects.create(
                org=mrf.org,
                candidate=candidate,
                mrf=mrf,
                mrf_line_item=mrf_li,
                site=mrf.site,
                job_role=mrf_li.job_role,
                current_stage=current_stage,
            )

            ApplicationStageHistory.objects.create(
                hiring_application=hiring_app,
                from_stage=None,
                to_stage=current_stage,
                from_status='',
                to_status=hiring_app.status,
                moved_by=user,
                comment='Application created via manual resume intake.',
            )

        return {
            'candidate': candidate,
            'resume': resume,
            'skills': skills_out,
            'hiring_application': hiring_app,
        }


def import_resume_file_for_role(
    *,
    user,
    uploaded_file,
    target_job_role,
    source_type: str = 'bulk_upload',
    view_only_note: str = '',
    import_batch_id: str = '',
    original_filename: str = '',
    content_type: str = '',
    size_bytes=None,
    stored_file_name: str = '',
    document_type: str = '',
    billing_type: str = None,
) -> dict:
    """
    Import one resume file when HR selected the target role before upload.
    The upload is parsed deterministically first, then attached to a candidate.
    """
    from .models import Candidate, Resume
    from .resume_parser.deterministic_parser import parse_resume_text
    from .resume_parser.extraction import extract_text_from_bytes
    from .resume_parser.normalization import normalize_parsed_json
    from .resume_parser.persistence import persist_parsed_data
    from .resume_parser.validation import validate_parsed_json

    org = user.org
    if target_job_role.org_id != org.id:
        raise ValidationError({'target_job_role': 'Target job role does not belong to your organization.'})

    filename = original_filename or getattr(uploaded_file, 'name', '') or 'resume'
    content_type = content_type or getattr(uploaded_file, 'content_type', '') or ''
    document_type = document_type or determine_document_type(filename, content_type)
    file_hash = compute_file_hash(uploaded_file)

    existing = Resume.objects.filter(
        candidate__org=org,
        file_hash=file_hash,
        status='indexed',
    ).select_related('candidate').first()
    if existing:
        return {
            'status': 'duplicate_file',
            'filename': filename,
            'candidate': existing.candidate,
            'resume': existing,
            'detail': 'An indexed resume with the same file already exists.',
        }

    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)

    try:
        raw_text, cleaned_text, extraction_engine, extraction_confidence = extract_text_from_bytes(
            raw_bytes,
            content_type=content_type,
            original_filename=filename,
        )
    except Exception as exc:
        raise ValidationError({'file': f'Could not extract text from {filename}: {exc}'})

    parsed_json = parse_resume_text(cleaned_text)
    validation_errors, missing_fields = validate_parsed_json(parsed_json)
    normalized = normalize_parsed_json(parsed_json)
    phone_normalized = normalized.get('phone_normalized') or ''
    if not phone_normalized:
        raise ValidationError({
            'phone': (
                f'No valid Indian mobile number found in {filename}. '
                'Upload through manual candidate intake for review.'
            )
        })

    confidence = parsed_json.get('confidence')

    with transaction.atomic():
        candidate, _ = _get_or_create_import_candidate(
            org=org,
            phone_normalized=phone_normalized,
            phone=normalized.get('phone') or phone_normalized,
            normalized=normalized,
            target_job_role=target_job_role,
            source='import_' if source_type in ('bulk_upload', 'excel_import') else 'qr',
            billing_type=billing_type,
        )

        resume = Resume.objects.create(
            candidate=candidate,
            file=stored_file_name or uploaded_file,
            original_filename=filename,
            content_type=content_type,
            size_bytes=size_bytes if size_bytes is not None else getattr(uploaded_file, 'size', 0),
            document_type=document_type,
            source_type=source_type,
            target_job_role=target_job_role,
            target_role_source=source_type if source_type in {'bulk_upload', 'campaign', 'qr_intake'} else 'bulk_upload',
            import_batch_id=import_batch_id,
            status='validating',
            uploaded_by=user,
            file_hash=file_hash,
            view_only_note=view_only_note,
            raw_text=raw_text,
            cleaned_text=cleaned_text,
            extraction_engine=extraction_engine,
            extraction_confidence=extraction_confidence,
            parser_engine='deterministic_v1',
        )

        persist_parsed_data(
            resume,
            normalized,
            parsed_json,
            validation_errors,
            missing_fields,
            confidence,
        )

        conf_dec = _decimal_confidence(confidence)
        Resume.objects.filter(pk=resume.pk).update(
            status='indexed',
            parser_confidence=conf_dec,
        )
        resume.status = 'indexed'
        resume.parser_confidence = conf_dec

    return {
        'status': 'indexed',
        'filename': filename,
        'candidate': candidate,
        'resume': resume,
        'detail': 'Resume imported and indexed.',
    }


def create_resume_import_batch(user, uploaded_files, target_job_role=None, source_type='bulk_upload', view_only_note='', billing_type=None):
    """Create a bulk import batch and enqueue one Celery task per file."""
    from .models import ResumeImportBatch, ResumeImportItem

    with transaction.atomic():
        batch = ResumeImportBatch.objects.create(
            org=user.org,
            target_job_role=target_job_role,
            source_type=source_type,
            status='queued',
            total_count=len(uploaded_files),
            view_only_note=view_only_note,
            created_by=user,
            billing_type=billing_type,
        )
        item_ids = []
        for uploaded_file in uploaded_files:
            item = ResumeImportItem.objects.create(
                batch=batch,
                file=uploaded_file,
                original_filename=getattr(uploaded_file, 'name', '') or 'resume',
                content_type=getattr(uploaded_file, 'content_type', '') or '',
                size_bytes=getattr(uploaded_file, 'size', 0),
                file_hash=compute_file_hash(uploaded_file),
                document_type=determine_document_type(
                    getattr(uploaded_file, 'name', ''),
                    getattr(uploaded_file, 'content_type', ''),
                ),
            )
            item_ids.append(item.pk)

        def _enqueue():
            from apps.talent.tasks import process_resume_import_item_task
            for item_id in item_ids:
                process_resume_import_item_task.delay(item_id)

        transaction.on_commit(_enqueue)

    return batch


def process_resume_import_item(item_id: int) -> None:
    """Process a single queued ResumeImportItem. Safe to retry."""
    from django.utils import timezone
    from .models import ResumeImportBatch, ResumeImportItem

    try:
        item = ResumeImportItem.objects.select_related(
            'batch',
            'batch__target_job_role',
            'batch__created_by',
        ).get(pk=item_id)
    except ResumeImportItem.DoesNotExist:
        return

    if item.status in ('indexed', 'duplicate_file'):
        return

    ResumeImportItem.objects.filter(pk=item.pk).update(status='processing', error_message='')
    ResumeImportBatch.objects.filter(pk=item.batch_id).update(status='processing')
    item.status = 'processing'

    try:
        item.file.open('rb')
        result = import_resume_file_for_role(
            user=item.batch.created_by,
            uploaded_file=item.file,
            target_job_role=item.batch.target_job_role,
            source_type=item.batch.source_type,
            view_only_note=item.batch.view_only_note,
            import_batch_id=str(item.batch_id),
            original_filename=item.original_filename,
            content_type=item.content_type,
            size_bytes=item.size_bytes,
            stored_file_name=item.file.name,
            document_type=item.document_type,
            billing_type=item.batch.billing_type,
        )
        item.file.close()

        status = result['status']
        ResumeImportItem.objects.filter(pk=item.pk).update(
            status=status,
            candidate=result.get('candidate'),
            resume=result.get('resume'),
            error_message='',
            processed_at=timezone.now(),
        )
    except Exception as exc:
        try:
            item.file.close()
        except Exception:
            pass
        ResumeImportItem.objects.filter(pk=item.pk).update(
            status='failed',
            error_message=_flatten_error(exc)[:2000],
            processed_at=timezone.now(),
        )

    _refresh_resume_import_batch_counts(item.batch_id)


def _refresh_resume_import_batch_counts(batch_id: int) -> None:
    from .models import ResumeImportBatch, ResumeImportItem

    items = ResumeImportItem.objects.filter(batch_id=batch_id)
    total = items.count()
    success = items.filter(status='indexed').count()
    duplicates = items.filter(status='duplicate_file').count()
    failed = items.filter(status='failed').count()
    manual_review = items.filter(status='manual_review').count()
    processed = success + duplicates + failed + manual_review

    if processed < total:
        status = 'processing'
    elif failed or manual_review:
        status = 'completed_with_errors'
    else:
        status = 'completed'

    ResumeImportBatch.objects.filter(pk=batch_id).update(
        status=status,
        total_count=total,
        processed_count=processed,
        success_count=success,
        duplicate_count=duplicates,
        failed_count=failed,
        manual_review_count=manual_review,
    )


def bulk_import_resume_files(user, files, target_job_role, source_type='bulk_upload', view_only_note='') -> dict:
    batch_id = uuid.uuid4().hex
    items = []
    created = 0
    duplicates = 0
    failed = 0

    for uploaded_file in files:
        filename = getattr(uploaded_file, 'name', '') or 'resume'
        try:
            result = import_resume_file_for_role(
                user=user,
                uploaded_file=uploaded_file,
                target_job_role=target_job_role,
                source_type=source_type,
                view_only_note=view_only_note,
                import_batch_id=batch_id,
            )
            if result['status'] == 'duplicate_file':
                duplicates += 1
            else:
                created += 1
            items.append(_resume_import_result_payload(result))
        except Exception as exc:
            failed += 1
            items.append({
                'filename': filename,
                'status': 'failed',
                'error': _flatten_error(exc),
            })

    return {
        'batch_id': batch_id,
        'created': created,
        'duplicates': duplicates,
        'failed': failed,
        'items': items,
    }


def _estimate_row_count(uploaded_file) -> int:
    filename = (getattr(uploaded_file, 'name', '') or '').lower()
    try:
        if filename.endswith('.csv'):
            uploaded_file.seek(0)
            content = uploaded_file.read()
            uploaded_file.seek(0)
            if isinstance(content, bytes):
                text = content.decode('utf-8-sig', errors='replace')
            else:
                text = content
            lines = [l for l in text.splitlines() if l.strip()]
            return max(0, len(lines) - 1)
        elif filename.endswith('.xlsx'):
            from openpyxl import load_workbook
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            total = 0
            for sheet in wb.worksheets:
                if sheet.max_row is not None and sheet.max_row > 1:
                    total += sheet.max_row - 1
            uploaded_file.seek(0)
            if total > 0:
                return total
    except Exception:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
    return 0


def _stream_candidate_sheet(uploaded_file):
    """Memory-efficient streaming generator for CSV and XLSX files."""
    filename = (getattr(uploaded_file, 'name', '') or '').lower()
    uploaded_file.seek(0)

    if filename.endswith('.csv'):
        content = uploaded_file.read()
        if isinstance(content, bytes):
            text = content.decode('utf-8-sig', errors='replace')
        else:
            text = content
        reader = csv.DictReader(StringIO(text))
        for index, row in enumerate(reader, start=2):
            normalized = {(_norm_header(k)): (v or '').strip() for k, v in row.items()}
            normalized['_source_row_number'] = index
            yield normalized
        return

    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValidationError({
                'file': 'openpyxl is not installed; install requirements to import .xlsx files.'
            })

        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            row_iterator = sheet.iter_rows(values_only=True)
            first_rows = []
            for _ in range(25):
                try:
                    r = next(row_iterator)
                    first_rows.append(r)
                except StopIteration:
                    break

            if not first_rows:
                continue

            header_index = _find_candidate_import_header_row(first_rows)
            if header_index is None:
                all_rows = first_rows
                for source_row_number, values in enumerate(all_rows, start=1):
                    if not values or not any(values):
                        continue
                    r_str = [str(v).strip() if v is not None else '' for v in values]
                    row = {
                        'name': r_str[0] if len(r_str) > 0 else '',
                        'phone': r_str[1] if len(r_str) > 1 else '',
                        'role': r_str[2] if len(r_str) > 2 else '',
                        '_source_row_number': source_row_number,
                    }
                    if row['name'] or row['phone']:
                        yield row
                source_row_number = len(all_rows)
                for values in row_iterator:
                    source_row_number += 1
                    if not values or not any(values):
                        continue
                    r_str = [str(v).strip() if v is not None else '' for v in values]
                    row = {
                        'name': r_str[0] if len(r_str) > 0 else '',
                        'phone': r_str[1] if len(r_str) > 1 else '',
                        'role': r_str[2] if len(r_str) > 2 else '',
                        '_source_row_number': source_row_number,
                    }
                    if row['name'] or row['phone']:
                        yield row
                continue

            headers = [_norm_header(h) for h in first_rows[header_index]]
            for sub_idx, values in enumerate(first_rows[header_index + 1:], start=header_index + 2):
                row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[idx] if idx < len(values) else ''
                    row[header] = '' if value is None else str(value).strip()
                if any(row.values()):
                    row['_source_row_number'] = sub_idx
                    yield row

            current_row_number = len(first_rows)
            for values in row_iterator:
                current_row_number += 1
                row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[idx] if idx < len(values) else ''
                    row[header] = '' if value is None else str(value).strip()
                if any(row.values()):
                    row['_source_row_number'] = current_row_number
                    yield row
        return

    raise ValidationError({'file': 'Upload a .csv or .xlsx file.'})


def _process_candidate_chunk(batch, chunk, org, default_target_job_role, billing_type, content_type, document_type):
    from apps.talent.models import Candidate, CandidateSkill, ResumeImportItem

    imported = 0
    duplicates = 0
    failed = 0
    import_items_to_create = []

    with transaction.atomic():
        for row in chunk:
            source_row_number = row.get('_source_row_number')
            try:
                role = _resolve_import_job_role(org, row, default_target_job_role)
                phone = _value(row, 'phone', 'mobile', 'contact')
                if not phone:
                    raise ValidationError({'phone': 'Phone/mobile is required.'})
                phone_normalized = normalize_phone(phone)
                first_name, last_name = _row_names(row)
                source_reference = f'excel_import_batch:{batch.pk}:row:{source_row_number}'
                candidate, created = Candidate.objects.get_or_create(
                    org=org,
                    phone_normalized=phone_normalized,
                    defaults={
                        'phone': phone,
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': _value(row, 'email') or '',
                        'current_location': _value(row, 'current_location', 'location') or '',
                        'total_experience_years': _decimal_or_none(
                            _value(row, 'total_experience_years', 'experience_years', 'experience')
                        ),
                        'current_company': _value(row, 'current_company', 'company') or '',
                        'current_role': _value(row, 'current_role', 'role', 'job_role') or (role.name if role else ''),
                        'target_job_role': role,
                        'source_reference': source_reference,
                        'source': 'import_',
                        'billing_type': billing_type,
                    },
                )
                if not created:
                    duplicates += 1
                    update_fields = []
                    if billing_type and candidate.billing_type != billing_type:
                        candidate.billing_type = billing_type
                        update_fields.append('billing_type')
                    if role and candidate.target_job_role != role:
                        candidate.target_job_role = role
                        update_fields.append('target_job_role')
                    if update_fields:
                        candidate.save(update_fields=update_fields)
                else:
                    imported += 1

                for skill_name in _split_skills(_value(row, 'skills', 'skill')):
                    normalized = normalize_skill_name(skill_name)
                    CandidateSkill.objects.get_or_create(
                        candidate=candidate,
                        normalized_skill_name=normalized,
                        defaults={
                            'skill_name': skill_name,
                            'source': 'excel_import',
                        },
                    )

                import_items_to_create.append(
                    ResumeImportItem(
                        batch=batch,
                        original_filename=f'Row {source_row_number}',
                        content_type=content_type,
                        size_bytes=0,
                        document_type=document_type,
                        row_number=source_row_number,
                        status='indexed',
                        candidate=candidate,
                    )
                )
            except Exception as exc:
                failed += 1
                error = _flatten_error(exc)
                import_items_to_create.append(
                    ResumeImportItem(
                        batch=batch,
                        original_filename=f'Row {source_row_number}',
                        content_type=content_type,
                        size_bytes=0,
                        document_type=document_type,
                        row_number=source_row_number,
                        status='failed',
                        error_message=error[:2000],
                    )
                )

        if import_items_to_create:
            ResumeImportItem.objects.bulk_create(import_items_to_create, batch_size=2000)

    return imported, duplicates, failed


def process_excel_import_batch(batch_id: int, chunk_size: int = 500) -> dict:
    """Celery task worker: processes Excel/CSV rows in memory-efficient chunks."""
    from apps.talent.models import ResumeImportBatch

    try:
        batch = ResumeImportBatch.objects.select_related('org', 'target_job_role', 'created_by').get(pk=batch_id)
    except ResumeImportBatch.DoesNotExist:
        return {'status': 'not_found'}

    if not batch.import_file:
        batch.status = 'failed'
        batch.save(update_fields=['status'])
        return {'status': 'no_file'}

    batch.status = 'processing'
    batch.save(update_fields=['status', 'updated_at'])

    imported = batch.success_count or 0
    duplicates = batch.duplicate_count or 0
    failed = batch.failed_count or 0
    processed = batch.processed_count or 0

    content_type = batch.content_type or ''
    document_type = batch.document_type or 'unknown'
    billing_type = batch.billing_type
    default_target_job_role = batch.target_job_role
    org = batch.org

    chunk = []

    try:
        batch.import_file.open('rb')
        stream = _stream_candidate_sheet(batch.import_file)

        for row in stream:
            chunk.append(row)
            if len(chunk) >= chunk_size:
                c_imported, c_dups, c_failed = _process_candidate_chunk(
                    batch=batch,
                    chunk=chunk,
                    org=org,
                    default_target_job_role=default_target_job_role,
                    billing_type=billing_type,
                    content_type=content_type,
                    document_type=document_type,
                )
                imported += c_imported
                duplicates += c_dups
                failed += c_failed
                processed += len(chunk)

                batch.processed_count = processed
                batch.success_count = imported
                batch.duplicate_count = duplicates
                batch.failed_count = failed
                if batch.total_count < processed:
                    batch.total_count = processed
                batch.save(update_fields=[
                    'processed_count', 'success_count', 'duplicate_count', 'failed_count', 'total_count', 'updated_at'
                ])
                chunk = []

        if chunk:
            c_imported, c_dups, c_failed = _process_candidate_chunk(
                batch=batch,
                chunk=chunk,
                org=org,
                default_target_job_role=default_target_job_role,
                billing_type=billing_type,
                content_type=content_type,
                document_type=document_type,
            )
            imported += c_imported
            duplicates += c_dups
            failed += c_failed
            processed += len(chunk)

        try:
            batch.import_file.close()
        except Exception:
            pass

        batch.total_count = processed
        batch.processed_count = processed
        batch.success_count = imported
        batch.duplicate_count = duplicates
        batch.failed_count = failed
        batch.status = 'completed_with_errors' if failed > 0 else 'completed'
        batch.save(update_fields=[
            'status', 'total_count', 'processed_count', 'success_count', 'duplicate_count', 'failed_count', 'updated_at'
        ])
    except Exception as exc:
        try:
            batch.import_file.close()
        except Exception:
            pass
        batch.status = 'failed'
        batch.save(update_fields=['status', 'updated_at'])
        return {'status': 'failed', 'error': str(exc)}

    return {
        'batch': batch.pk,
        'batch_id': batch.pk,
        'status': batch.status,
        'imported': imported,
        'duplicates': duplicates,
        'failed': failed,
        'total': processed,
    }


def import_candidates_from_excel(user, uploaded_file, default_target_job_role=None, source_type='excel_import', billing_type=None) -> dict:
    """Create an Excel/CSV import batch and enqueue Celery task for async chunk processing."""
    from apps.talent.models import ResumeImportBatch

    filename = getattr(uploaded_file, 'name', '') or 'candidate-import'
    content_type = getattr(uploaded_file, 'content_type', '') or ''
    size_bytes = getattr(uploaded_file, 'size', 0)
    document_type = determine_document_type(filename, content_type)

    uploaded_file.seek(0)
    total_count = _estimate_row_count(uploaded_file)
    uploaded_file.seek(0)

    with transaction.atomic():
        batch = ResumeImportBatch.objects.create(
            org=user.org,
            target_job_role=default_target_job_role,
            source_type=source_type,
            status='queued',
            total_count=total_count,
            import_file=uploaded_file,
            original_filename=filename,
            content_type=content_type,
            size_bytes=size_bytes,
            document_type=document_type,
            created_by=user,
            billing_type=billing_type,
        )

        def _enqueue():
            from apps.talent.tasks import process_excel_import_batch_task
            process_excel_import_batch_task.delay(batch.pk, chunk_size=500)

        transaction.on_commit(_enqueue)

    batch.refresh_from_db()
    items = []
    if batch.status in ('completed', 'completed_with_errors'):
        for item in batch.items.select_related('candidate').all():
            items.append({
                'row': item.row_number,
                'status': 'created' if item.status == 'indexed' else item.status,
                'candidate': item.candidate_id,
                'error': item.error_message,
                'target_job_role': default_target_job_role.pk if default_target_job_role else None,
            })

    return {
        'batch': batch.pk,
        'batch_id': batch.pk,
        'id': batch.pk,
        'status': batch.status,
        'document_type': document_type,
        'imported': batch.success_count,
        'duplicates': batch.duplicate_count,
        'failed': batch.failed_count,
        'total_count': batch.total_count,
        'processed_count': batch.processed_count,
        'message': 'Import started',
        'items': items,
    }


def _get_or_create_import_candidate(*, org, phone_normalized, phone, normalized, target_job_role, source, billing_type=None):
    from .models import Candidate

    first_name = normalized.get('first_name') or 'Unknown'
    last_name = normalized.get('last_name') or 'Unknown'
    defaults = {
        'phone': phone,
        'first_name': first_name,
        'middle_name': normalized.get('middle_name') or '',
        'last_name': last_name,
        'email': normalized.get('email') or '',
        'current_role': normalized.get('current_role') or target_job_role.name,
        'current_location': normalized.get('current_location') or '',
        'total_experience_years': _decimal_or_none(normalized.get('total_experience_years')),
        'current_company': normalized.get('current_company') or '',
        'target_job_role': target_job_role,
        'source': source,
        'billing_type': billing_type,
    }
    candidate, created = Candidate.objects.get_or_create(
        org=org,
        phone_normalized=phone_normalized,
        defaults=defaults,
    )
    if created:
        return candidate, created

    update_fields = []
    for field, incoming in defaults.items():
        if incoming is None or incoming == '':
            continue
        if field in ('first_name', 'last_name') and getattr(candidate, field) not in ('', 'Unknown'):
            continue
        if field == 'target_job_role' and candidate.target_job_role_id:
            continue
        if getattr(candidate, field) != incoming:
            setattr(candidate, field, incoming)
            update_fields.append(field)
    if update_fields:
        candidate.save(update_fields=list(dict.fromkeys(update_fields)))
    return candidate, created


def _decimal_confidence(confidence):
    if confidence is None:
        return None
    try:
        conf_dec = Decimal(str(confidence)).quantize(Decimal('0.01'))
        return max(Decimal('0.00'), min(Decimal('1.00'), conf_dec))
    except InvalidOperation:
        return None


def _resume_import_result_payload(result: dict) -> dict:
    candidate = result.get('candidate')
    resume = result.get('resume')
    return {
        'filename': result.get('filename'),
        'status': result.get('status'),
        'detail': result.get('detail'),
        'candidate': candidate.pk if candidate else None,
        'candidate_name': candidate.full_name if candidate else '',
        'resume': resume.pk if resume else None,
    }


def _flatten_error(exc) -> str:
    detail = getattr(exc, 'detail', None)
    if detail is not None:
        return str(detail)
    return str(exc)


def candidate_profile_quality(candidate) -> dict:
    """Compute a lightweight completeness score for search and review UX."""
    resume_count = getattr(candidate, 'profile_resume_count', None)
    if resume_count is None:
        resume_count = candidate.resumes.count()

    skill_count = getattr(candidate, 'profile_skill_count', None)
    if skill_count is None:
        skill_count = candidate.skills.count()

    experience_count = getattr(candidate, 'profile_experience_count', None)
    if experience_count is None:
        experience_count = candidate.experiences.count()

    education_count = getattr(candidate, 'profile_education_count', None)
    if education_count is None:
        education_count = candidate.educations.count()

    checks = {
        'phone_present': bool(candidate.phone_normalized or candidate.phone),
        'mapped_role_present': bool(candidate.target_job_role_id),
        'resume_file_present': resume_count > 0,
        'skills_present': skill_count > 0,
        'experience_present': (
            candidate.total_experience_years is not None
            or experience_count > 0
        ),
        'education_present': education_count > 0,
        'location_present': bool(candidate.current_location or candidate.preferred_location),
    }
    weights = {
        'phone_present': 20,
        'mapped_role_present': 20,
        'resume_file_present': 15,
        'skills_present': 15,
        'experience_present': 15,
        'education_present': 5,
        'location_present': 10,
    }
    return {
        'score': sum(weights[key] for key, ok in checks.items() if ok),
        'checks': checks,
        'missing': [key for key, ok in checks.items() if not ok],
    }


JOURNEY_STATUS_LABELS = {
    'available': 'Available',
    'available_from_date': 'Available from date',
    'notice_period': 'Notice period',
    'not_available': 'Not available',
    'unknown': 'Unknown',
    'shortlisted': 'Shortlisted',
    'sent_to_client': 'Sent to client',
    'client_approved': 'Client approved',
    'client_rejected': 'Client rejected',
    'interview': 'Interview / verification',
    'offer_draft': 'Offer draft',
    'offer_released': 'Offer released',
    'offer_accepted': 'Offer accepted',
    'offer_declined': 'Offer declined',
    'offer_withdrawn': 'Offer withdrawn',
    'offer_expired': 'Offer expired',
    'deployed': 'Deployed',
    'deployment_planned': 'Deployment planned',
    'deployment_completed': 'Deployment completed',
    'deployment_transferred': 'Deployment transferred',
    'deployment_cancelled': 'Deployment cancelled',
    'employee_active': 'Employee active',
    'employee_inactive': 'Employee inactive',
    'employee_suspended': 'Employee suspended',
    'exited': 'Exited',
    'rejected': 'Rejected',
    'cancelled': 'Cancelled',
    'duplicate': 'Duplicate',
    'blacklisted': 'Blacklisted',
    'do_not_contact': 'Do not contact',
}


def _journey_payload(
    *,
    status,
    latest_application=None,
    latest_offer_status=None,
    employee=None,
    deployment=None,
):
    return {
        'journey_status': status,
        'journey_status_label': JOURNEY_STATUS_LABELS.get(status, status.replace('_', ' ').title()),
        'latest_application_id': latest_application.pk if latest_application else None,
        'latest_application_status': latest_application.status if latest_application else None,
        'latest_offer_status': latest_offer_status,
        'employee_id': employee.pk if employee else None,
        'employee_status': employee.status if employee else None,
        'deployment_id': deployment.pk if deployment else None,
        'deployment_status': deployment.status if deployment else None,
    }


def _status_from_offer(offer_status):
    return {
        'draft': 'offer_draft',
        'released': 'offer_released',
        'accepted': 'offer_accepted',
        'declined': 'offer_declined',
        'withdrawn': 'offer_withdrawn',
        'expired': 'offer_expired',
    }.get(offer_status)


def _status_from_application(application):
    if application is None:
        return None

    try:
        offer_status = application.offer.status
    except Exception:
        offer_status = None

    offer_journey = _status_from_offer(offer_status)
    if offer_journey:
        return offer_journey, offer_status

    app_status = application.status
    if app_status == 'client_review':
        if application.client_decision == 'approved':
            return 'client_approved', offer_status
        if application.client_decision == 'rejected':
            return 'client_rejected', offer_status
        return 'sent_to_client', offer_status
    if app_status in ('interview_scheduled', 'interview_in_progress'):
        return 'interview', offer_status
    if app_status == 'selected':
        return 'client_approved', offer_status
    if app_status == 'offer_released':
        return 'offer_released', offer_status
    if app_status == 'offer_accepted':
        return 'offer_accepted', offer_status
    if app_status == 'offer_declined':
        return 'offer_declined', offer_status
    if app_status == 'deployed':
        return 'deployed', offer_status
    if app_status in ('shortlisted', 'rejected', 'cancelled'):
        return app_status, offer_status
    if app_status == 'draft':
        return 'shortlisted', offer_status
    return app_status or None, offer_status


def candidate_journey_status(candidate) -> dict:
    """
    Return the backend-derived business journey for a candidate.

    This is intentionally derived at read time from source-of-truth records:
    Employee/SiteDeployment first, then latest HiringApplication/Offer, then
    candidate lifecycle and availability. The frontend should display this
    payload, not recreate this decision tree.
    """
    if candidate.is_blacklisted or candidate.lifecycle_status == 'blacklisted':
        return _journey_payload(status='blacklisted')
    if candidate.do_not_contact or candidate.lifecycle_status == 'do_not_contact':
        return _journey_payload(status='do_not_contact')
    if candidate.is_duplicate or candidate.lifecycle_status == 'duplicate':
        return _journey_payload(status='duplicate')

    employee = (
        candidate.employee_records
        .prefetch_related('deployments')
        .order_by('-updated_at', '-id')
        .first()
    )
    if employee:
        deployment = (
            employee.deployments
            .order_by('-start_date', '-updated_at', '-id')
            .first()
        )
        active_deployment = (
            employee.deployments
            .filter(status='active')
            .order_by('-start_date', '-updated_at', '-id')
            .first()
        )
        if employee.status == 'exited':
            return _journey_payload(
                status='exited',
                employee=employee,
                deployment=deployment,
            )
        if employee.status == 'suspended':
            return _journey_payload(
                status='employee_suspended',
                employee=employee,
                deployment=active_deployment or deployment,
            )
        if active_deployment:
            return _journey_payload(
                status='deployed',
                employee=employee,
                deployment=active_deployment,
            )
        if deployment and deployment.status in ('planned', 'completed', 'transferred', 'cancelled'):
            return _journey_payload(
                status=f'deployment_{deployment.status}',
                employee=employee,
                deployment=deployment,
            )
        return _journey_payload(
            status=f'employee_{employee.status}',
            employee=employee,
            deployment=deployment,
        )

    latest_application = (
        candidate.hiring_applications
        .select_related('offer')
        .order_by('-updated_at', '-created_at', '-id')
        .first()
    )
    app_status, offer_status = _status_from_application(latest_application) if latest_application else (None, None)
    if app_status:
        return _journey_payload(
            status=app_status,
            latest_application=latest_application,
            latest_offer_status=offer_status,
        )

    if candidate.availability_status in ('available_from_date', 'notice_period', 'not_available', 'unknown'):
        return _journey_payload(status=candidate.availability_status)
    if candidate.availability_status == 'currently_deployed':
        return _journey_payload(status='deployed')
    return _journey_payload(status='available')


def _read_candidate_sheet(uploaded_file) -> list[dict]:
    filename = (getattr(uploaded_file, 'name', '') or '').lower()
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    uploaded_file.seek(0)

    if filename.endswith('.csv'):
        text = raw.decode('utf-8-sig', errors='replace')
        output = []
        for index, row in enumerate(csv.DictReader(StringIO(text)), start=2):
            normalized = {(_norm_header(k)): (v or '').strip() for k, v in row.items()}
            normalized['_source_row_number'] = index
            output.append(normalized)
        return output

    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValidationError({
                'file': 'openpyxl is not installed; install requirements to import .xlsx files.'
            })
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        output = []

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_index = _find_candidate_import_header_row(rows)
            if header_index is None:
                for source_row_number, values in enumerate(rows, start=1):
                    if not values or not any(values):
                        continue
                    r_str = [str(v).strip() if v is not None else '' for v in values]
                    row = {
                        'name': r_str[0] if len(r_str) > 0 else '',
                        'phone': r_str[1] if len(r_str) > 1 else '',
                        'role': r_str[2] if len(r_str) > 2 else '',
                        '_source_row_number': source_row_number,
                    }
                    if row['name'] or row['phone']:
                        output.append(row)
                continue

            headers = [_norm_header(h) for h in rows[header_index]]
            for source_row_number, values in enumerate(rows[header_index + 1:], start=header_index + 2):
                row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[idx] if idx < len(values) else ''
                    row[header] = '' if value is None else str(value).strip()
                if any(row.values()):
                    row['_source_row_number'] = source_row_number
                    output.append(row)
        return output

    raise ValidationError({'file': 'Upload a .csv or .xlsx file.'})


def _norm_header(value) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def _find_candidate_import_header_row(rows) -> int | None:
    """
    Locate the actual candidate header row in an XLSX sheet.

    Business users often add a title row above the table.  The importer needs to
    tolerate that and start at the row containing phone/name columns.
    """
    phone_headers = {'phone', 'mobile', 'contact', 'mobile_number', 'phone_number', 'contact_number'}
    name_headers = {'first_name', 'firstname', 'last_name', 'lastname', 'full_name', 'name'}
    profile_headers = {
        'email', 'current_role', 'role', 'job_role', 'current_location', 'location',
        'total_experience_years', 'experience_years', 'experience', 'skills',
    }

    for idx, values in enumerate(rows[:20]):
        headers = {_norm_header(value) for value in values if _norm_header(value)}
        if not headers:
            continue
        has_phone = bool(headers & phone_headers)
        has_name = bool(headers & name_headers)
        has_profile = bool(headers & profile_headers)
        if has_phone and (has_name or has_profile):
            return idx
    return None


def _value(row: dict, *keys: str) -> str:
    for key in keys:
        val = row.get(_norm_header(key))
        if val not in (None, ''):
            return str(val).strip()
    return ''


def _row_names(row: dict) -> tuple[str, str]:
    first = _value(row, 'first_name', 'firstname')
    last = _value(row, 'last_name', 'lastname')
    full = _value(row, 'full_name', 'name')
    if full and not first:
        parts = full.split()
        first = parts[0] if parts else 'Unknown'
        last = ' '.join(parts[1:]) if len(parts) > 1 else 'Unknown'
    return first or 'Unknown', last or 'Unknown'


def _resolve_import_job_role(org, row: dict, default_role):
    from apps.jobs.models import JobRole

    raw = _value(row, 'target_job_role', 'job_role', 'role', 'mapped_role')
    if not raw:
        return default_role
    normalized = raw.strip().lower()
    role = JobRole.objects.filter(org=org, code__iexact=normalized).first()
    if role:
        return role
    return JobRole.objects.filter(org=org, name__iexact=raw.strip()).first() or default_role


def _decimal_or_none(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _split_skills(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r'[,;|]', value) if part.strip()]


# ─── Review services ──────────────────────────────────────────────────────────

def apply_review_service(resume, user, validated_data: dict):
    """
    Apply HR corrections to a resume: update candidate, replace parsed data,
    set status=indexed, create TalentResumeReview audit record.
    Returns the created TalentResumeReview.
    """
    from .models import (
        Candidate, CandidateSkill, CandidateExperience,
        CandidateEducation, ParsedResume, TalentResumeReview,
    )

    with transaction.atomic():
        candidate = resume.candidate
        previous_status = resume.status

        # 1. Update candidate fields (never overwrite with blank)
        candidate_data = validated_data.get('candidate') or {}
        if candidate_data:
            update_fields = []

            phone_val = (candidate_data.get('phone') or '').strip()
            if phone_val:
                phone_normalized = normalize_phone(phone_val)
                if candidate.phone != phone_val:
                    candidate.phone = phone_val
                    candidate.phone_normalized = phone_normalized
                    update_fields.extend(['phone', 'phone_normalized'])

            for field in [
                'first_name', 'middle_name', 'last_name', 'email',
                'current_role', 'current_company', 'current_location',
                'total_experience_years', 'expected_ctc', 'current_ctc',
                'notice_period_days',
            ]:
                val = candidate_data.get(field)
                if val is None:
                    continue
                if isinstance(val, str) and not val.strip():
                    continue
                if field == 'total_experience_years':
                    try:
                        val = Decimal(str(val))
                    except InvalidOperation:
                        continue
                if getattr(candidate, field, None) != val:
                    setattr(candidate, field, val)
                    update_fields.append(field)

            if update_fields:
                candidate.save(update_fields=list(dict.fromkeys(update_fields)))

        # 2. Replace parsed/reviewed skills — manual skills untouched
        skills_data = validated_data.get('skills')
        if skills_data is not None:
            CandidateSkill.objects.filter(
                candidate=candidate,
                source_resume=resume,
                source__in=['parsed', 'reviewed'],
            ).delete()
            for skill_d in skills_data:
                name = skill_d['skill_name'].strip()
                if not name:
                    continue
                CandidateSkill.objects.create(
                    candidate=candidate,
                    skill_name=name,
                    normalized_skill_name=name.lower(),
                    years_experience=skill_d.get('years_experience'),
                    proficiency=skill_d.get('proficiency') or '',
                    source='reviewed',
                    source_resume=resume,
                )

        # 3. Replace experience for this resume
        experience_data = validated_data.get('experience')
        if experience_data is not None:
            CandidateExperience.objects.filter(candidate=candidate, source_resume=resume).delete()
            for exp_d in experience_data:
                CandidateExperience.objects.create(
                    candidate=candidate,
                    source_resume=resume,
                    job_title=exp_d.get('job_title') or '',
                    company_name=exp_d.get('company_name') or '',
                    industry=exp_d.get('industry') or '',
                    start_date=exp_d.get('start_date'),
                    end_date=exp_d.get('end_date'),
                    is_current=bool(exp_d.get('is_current', False)),
                    duration_months=exp_d.get('duration_months'),
                    description=exp_d.get('description') or '',
                    responsibilities=exp_d.get('responsibilities') or [],
                )

        # 4. Replace education for this resume
        education_data = validated_data.get('education')
        if education_data is not None:
            CandidateEducation.objects.filter(candidate=candidate, source_resume=resume).delete()
            for edu_d in education_data:
                CandidateEducation.objects.create(
                    candidate=candidate,
                    source_resume=resume,
                    degree=edu_d.get('degree') or '',
                    specialization=edu_d.get('specialization') or '',
                    institute=edu_d.get('institute') or '',
                    start_year=edu_d.get('start_year'),
                    end_year=edu_d.get('end_year'),
                )

        # 5. Upsert ParsedResume — clear errors, mark confidence 1.0
        corrected_normalized = {}
        if candidate_data:
            corrected_normalized.update({
                k: v for k, v in candidate_data.items() if k != 'phone'
            })
        if skills_data is not None:
            corrected_normalized['skills'] = [
                {
                    'name': s['skill_name'],
                    'normalized_name': s['skill_name'].lower(),
                    'years_experience': str(s['years_experience']) if s.get('years_experience') is not None else None,
                    'proficiency': s.get('proficiency') or '',
                }
                for s in skills_data
            ]
        if experience_data is not None:
            corrected_normalized['experience'] = [
                {k: str(v) if hasattr(v, 'isoformat') else v for k, v in e.items()}
                for e in experience_data
            ]
        if education_data is not None:
            corrected_normalized['education'] = list(education_data)

        ParsedResume.objects.update_or_create(
            resume=resume,
            defaults={
                'normalized_json': corrected_normalized,
                'validation_errors': [],
                'missing_fields': [],
                'confidence': Decimal('1.00'),
            },
        )

        # 6. Update resume to indexed
        Resume.objects.filter(pk=resume.pk).update(
            status='indexed',
            manual_review_reason='',
            error_message='',
            parser_confidence=Decimal('1.00'),
        )
        resume.status = 'indexed'

        # 7. Create audit record
        review = TalentResumeReview.objects.create(
            org=candidate.org,
            resume=resume,
            candidate=candidate,
            reviewed_by=user,
            review_type='correction',
            previous_status=previous_status,
            new_status='indexed',
            review_note=validated_data.get('review_note', ''),
            correction_payload={
                k: v for k, v in validated_data.items() if k != 'review_note'
            },
        )

    return review


def resolve_duplicate_service(resume, user, validated_data: dict):
    """
    Resolve a duplicate candidate/resume situation.
    Returns the created TalentResumeReview.
    """
    from .models import Candidate, TalentResumeReview

    resolution = validated_data['resolution']
    existing_candidate = validated_data.get('candidate')
    note = validated_data.get('note', '')

    with transaction.atomic():
        candidate = resume.candidate
        previous_status = resume.status

        if resolution == 'link_existing':
            if existing_candidate is None:
                raise ValidationError({'candidate': 'candidate is required for link_existing resolution.'})
            if existing_candidate.org_id != candidate.org_id:
                raise ValidationError({'candidate': 'Target candidate belongs to a different organisation.'})
            resume.candidate = existing_candidate
            resume.save(update_fields=['candidate'])
            Resume.objects.filter(pk=resume.pk).update(
                status='manual_review',
                manual_review_reason='Linked to existing candidate after duplicate review.',
            )
            resume.status = 'manual_review'
            audit_candidate = existing_candidate

        elif resolution == 'mark_duplicate':
            update_c_fields = ['is_duplicate']
            candidate.is_duplicate = True
            if existing_candidate:
                if existing_candidate.org_id != candidate.org_id:
                    raise ValidationError({'candidate': 'Target candidate belongs to a different organisation.'})
                candidate.duplicate_of = existing_candidate
                update_c_fields.append('duplicate_of')
            candidate.save(update_fields=update_c_fields)
            Resume.objects.filter(pk=resume.pk).update(status='duplicate_file')
            resume.status = 'duplicate_file'
            audit_candidate = candidate

        elif resolution == 'keep_separate':
            candidate.is_duplicate = False
            candidate.duplicate_of = None
            candidate.save(update_fields=['is_duplicate', 'duplicate_of'])
            if resume.status == 'duplicate_file':
                Resume.objects.filter(pk=resume.pk).update(
                    status='manual_review',
                    manual_review_reason='Kept as separate candidate after duplicate review.',
                )
                resume.status = 'manual_review'
            audit_candidate = candidate

        else:
            raise ValidationError({'resolution': f'Unknown resolution: {resolution}'})

        review = TalentResumeReview.objects.create(
            org=audit_candidate.org,
            resume=resume,
            candidate=audit_candidate,
            reviewed_by=user,
            review_type='duplicate_resolution',
            previous_status=previous_status,
            new_status=resume.status,
            review_note=note,
            correction_payload={
                'resolution': resolution,
                'candidate': existing_candidate.pk if existing_candidate else None,
            },
        )

    return review


def merge_candidate_service(source_candidate, target_candidate, user, note='') -> dict:
    """
    Merge a duplicate candidate into a target candidate.

    Hiring applications are protected and remain on the source candidate. The
    service refuses those merges so hiring/deployment history is not silently
    rewritten.
    """
    from apps.hiring.models import HiringApplication, CandidateMatchResult
    from .models import (
        CandidateSkill, CandidateExperience, CandidateEducation,
        Resume, ResumeImportItem, TalentResumeReview,
    )

    if source_candidate.pk == target_candidate.pk:
        raise ValidationError({'target_candidate': 'Choose a different target candidate.'})
    if source_candidate.org_id != target_candidate.org_id:
        raise ValidationError({'target_candidate': 'Target candidate belongs to a different organisation.'})
    if HiringApplication.objects.filter(candidate=source_candidate).exists():
        raise ValidationError({
            'source_candidate': (
                'Cannot merge a candidate linked to hiring applications. '
                'Resolve hiring/deployment history first.'
            )
        })

    with transaction.atomic():
        Resume.objects.filter(candidate=source_candidate).update(candidate=target_candidate)
        ResumeImportItem.objects.filter(candidate=source_candidate).update(candidate=target_candidate)
        CandidateExperience.objects.filter(candidate=source_candidate).update(candidate=target_candidate)
        CandidateEducation.objects.filter(candidate=source_candidate).update(candidate=target_candidate)
        CandidateMatchResult.objects.filter(candidate=source_candidate).update(candidate=target_candidate)

        for skill in CandidateSkill.objects.filter(candidate=source_candidate):
            normalized = skill.normalized_skill_name or normalize_skill_name(skill.skill_name)
            existing = CandidateSkill.objects.filter(
                candidate=target_candidate,
                normalized_skill_name=normalized,
            ).first()
            if existing:
                skill.delete()
                continue
            skill.candidate = target_candidate
            skill.normalized_skill_name = normalized
            skill.save(update_fields=['candidate', 'normalized_skill_name'])

        update_fields = []
        for field in [
            'email', 'current_location', 'total_experience_years',
            'current_ctc', 'expected_ctc', 'preferred_location',
            'notice_period_days', 'current_company', 'current_role',
            'target_job_role', 'source_reference',
        ]:
            source_value = getattr(source_candidate, field)
            target_value = getattr(target_candidate, field)
            if target_value in (None, '') and source_value not in (None, ''):
                setattr(target_candidate, field, source_value)
                update_fields.append(field)
        if update_fields:
            target_candidate.save(update_fields=update_fields)

        source_candidate.is_duplicate = True
        source_candidate.duplicate_of = target_candidate
        source_candidate.lifecycle_status = 'duplicate'
        source_candidate.save(update_fields=['is_duplicate', 'duplicate_of', 'lifecycle_status'])

        review_resume = target_candidate.resumes.order_by('-uploaded_at').first()
        if review_resume:
            TalentResumeReview.objects.create(
                org=target_candidate.org,
                resume=review_resume,
                candidate=target_candidate,
                reviewed_by=user,
                review_type='candidate_merge',
                previous_status='',
                new_status='',
                review_note=note,
                correction_payload={
                    'source_candidate': source_candidate.pk,
                    'target_candidate': target_candidate.pk,
                },
            )

    return {
        'source_candidate': source_candidate,
        'target_candidate': target_candidate,
        'profile_quality': candidate_profile_quality(target_candidate),
    }
